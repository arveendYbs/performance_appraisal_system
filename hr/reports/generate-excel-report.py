#!/usr/bin/env python3
"""
Generate Appraisal Excel Report - SIMPLIFIED VERSION
Creates valid Excel files compatible with Microsoft Excel
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_report(data_file, output_file):
    try:
        # Load data
        with open(data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        employee = data['employee']
        appraisals = data['appraisals']
        year = data['year']
        
        print(f"Processing report for: {employee['name']}")
        print(f"Number of appraisals: {len(appraisals)}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{year} Report"
        
        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        subheader_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
        subheader_font = Font(bold=True, size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # Title
        ws['A1'] = f"Summary of {year} Appraisal Ratings - {employee.get('company_name', 'N/A')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:R1')
        current_row = 3
        
        # Section headers
        ws.merge_cells(f'J{current_row}:V{current_row}')
        ws[f'J{current_row}'] = 'Performance Assessment - Employee Scores'
        ws[f'J{current_row}'].fill = subheader_fill
        ws[f'J{current_row}'].font = subheader_font
        ws[f'J{current_row}'].alignment = center_align
        
        ws.merge_cells(f'W{current_row}:AI{current_row}')
        ws[f'W{current_row}'] = 'Performance Assessment - Manager Scores'
        ws[f'W{current_row}'].fill = subheader_fill
        ws[f'W{current_row}'].font = subheader_font
        ws[f'W{current_row}'].alignment = center_align
        
        current_row += 1
        
        # Column headers
        headers = [
            'Company', 'Dept', 'Name', 'Staff No.', 'Form', 'Role', 
            'Position', 'Date Joined', 'Period'
        ]
        
        # Add 12 section columns for employee
        for i in range(1, 13):
            headers.append(f'S{i}')
        
        headers.extend(['Total', 'Score', 'Rating'])
        
        # Add 12 section columns for manager
        for i in range(1, 13):
            headers.append(f'S{i}')
        
        headers.extend(['Total', 'Score', 'Final Rating'])
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        current_row += 1
        
        # Set column widths
        widths = {
            'A': 15, 'B': 15, 'C': 20, 'D': 12, 'E': 18, 'F': 12,
            'G': 20, 'H': 12, 'I': 15
        }
        # J to U (employee sections) - 8 width each
        for col in range(ord('J'), ord('V')):
            widths[chr(col)] = 8
        # V, W, X (totals/scores) - 10 width
        widths['V'] = 10
        widths['W'] = 10
        widths['X'] = 10
        # Y to AJ (manager sections) - 8 width each
        for col_num in range(25, 37):  # Y=25 to AJ=36
            widths[get_column_letter(col_num)] = 8
        # Totals/scores
        widths['AK'] = 10
        widths['AL'] = 10
        widths['AM'] = 12
        
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Write data for each appraisal
        for appraisal in appraisals:
            sections = appraisal.get('sections', [])
            
            # Basic info
            row_data = [
                employee.get('company_name', ''),
                employee.get('department', ''),
                employee.get('name', ''),
                employee.get('emp_number', ''),
                appraisal.get('form_title', ''),
                employee.get('role', '').title(),
                employee.get('position', ''),
                employee.get('date_joined', ''),
                f"{appraisal.get('period_from', '')} to {appraisal.get('period_to', '')}"
            ]
            
            # Employee section scores (up to 12)
            for i in range(12):
                if i < len(sections):
                    row_data.append(sections[i].get('employee_score', 0))
                else:
                    row_data.append('')
            
            # Employee total (formula)
            row_data.append(f'=SUM(J{current_row}:U{current_row})')
            
            # Employee calculated score (formula)
            role = employee.get('role', 'employee').lower()
            if 'manager' in role or 'admin' in role:
                divisor = '1.2'
            elif 'worker' in role:
                divisor = '0.8'
            else:
                divisor = '1'
            row_data.append(f'=ROUND(V{current_row}/{divisor},0)')
            
            # Employee rating (formula)
            rating_formula = f'=IF(W{current_row}=0,"",IF(W{current_row}<50,"C",IF(W{current_row}<60,"B-",IF(W{current_row}<75,"B",IF(W{current_row}<85,"B+","A")))))'
            row_data.append(rating_formula)
            
            # Manager section scores (up to 12)
            for i in range(12):
                if i < len(sections):
                    row_data.append(sections[i].get('manager_score', 0))
                else:
                    row_data.append('')
            
            # Manager total (formula)
            row_data.append(f'=SUM(Y{current_row}:AJ{current_row})')
            
            # Manager calculated score (formula)
            row_data.append(f'=ROUND(AK{current_row}/{divisor},0)')
            
            # Final rating (formula)
            final_rating_formula = f'=IF(AL{current_row}=0,"",IF(AL{current_row}<50,"C",IF(AL{current_row}<60,"B-",IF(AL{current_row}<75,"B",IF(AL{current_row}<85,"B+","A")))))'
            row_data.append(final_rating_formula)
            
            # Write row
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.border = thin_border
                
                # Alignment
                if col_idx <= 9:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
            
            current_row += 1
        
        # Freeze panes
        ws.freeze_panes = 'A5'
        
        # Save workbook
        wb.save(output_file)
        print(f"Excel report generated: {output_file}")
        
        # Verify file was created
        import os
        if os.path.exists(output_file):
            file_size = os.path.getsize(output_file)
            print(f"File size: {file_size} bytes")
            if file_size > 0:
                return True
            else:
                print("ERROR: File is empty!")
                return False
        else:
            print("ERROR: File was not created!")
            return False
            
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python generate-excel-report.py <data_json> <output_xlsx>")
        sys.exit(1)
    
    data_file = sys.argv[1]
    output_file = sys.argv[2]
    
    print(f"Input JSON: {data_file}")
    print(f"Output Excel: {output_file}")
    
    success = create_report(data_file, output_file)
    sys.exit(0 if success else 1)